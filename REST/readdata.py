import openpyxl

wb=openpyxl.load_workbook('C:/Users/Dusa/PycharmProjects/API/REST/data.xlsx')
ws=wb['Sheet1']
row=ws.max_row
#print(row)
for i in range(2,row+1):
    first_name=ws.cell(i,1).value
    print(first_name)
