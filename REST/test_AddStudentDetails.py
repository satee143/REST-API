import pytest
import json
import jsonpath
import requests
import openpyxl


def test_Add_Student_Details():
    uri='http://thetestingworldapi.com/api/studentsDetails'

    f=open('C:/Users/Dusa/PycharmProjects/API/REST/data.json','r')
    data=f.read()

    wb = openpyxl.load_workbook('C:/Users/Dusa/PycharmProjects/API/REST/data.xlsx')
    ws = wb['Sheet1']
    row = ws.max_row
    # print(row)
    jsonformat = json.loads(data)
    for i in range(2, row + 1):
        jsonformat['first_name']=str(ws.cell(i, 1).value)
        jsonformat['middle_name'] = str(ws.cell(i, 2).value)
        jsonformat['last_name'] = str(ws.cell(i, 3).value)
        jsonformat['date_of_birth'] = str(ws.cell(i, 4).value)
        res=requests.post(uri,jsonformat)
        print(res.status_code)
        jsone=json.loads(res.content)
        print(jsonpath.jsonpath(jsone,'id'))


